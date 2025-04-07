from . import teachers, subjects, audiences
from openpyxl import load_workbook


def create_workbook(schedule):
    print("Создаём Excel документ для кастомизации расписания...")

    file_path = 'assets/template.xlsx'
    workbook = load_workbook(file_path)
    
    my_subjects = subjects.get_uniq_subjects(schedule)
    fill_subjects_sheet(workbook, my_subjects)

    my_teachers = teachers.get_uniq_teachers(schedule)
    fill_teachers_sheet(workbook, my_teachers)

    my_audiences = audiences.get_uniq_audiences(schedule)
    fill_audiences_sheet(workbook, my_audiences)

    fill_lessons_sheet(workbook, schedule)

    output_file_path = 'output/schedule.xlsx'
    workbook.save(output_file_path)
    
    print("Расписание успешно создано и сохранено в output/schedule.xlsx")


def fill_subjects_sheet(workbook, my_subjects):
    subjects_sheet = workbook['Subjects']
    row = 2
    for subject in my_subjects:
        subjects_sheet[f'A{row}'] = subject
        row += 1


def fill_teachers_sheet(workbook, my_teachers):
    teachers_sheet = workbook['Teachers']
    row = 2
    for teacher in my_teachers:
        teachers_sheet[f'A{row}'] = teacher['name']
        teachers_sheet[f'B{row}'] = teacher['full_name']
        row += 1


def fill_audiences_sheet(workbook, my_audiences):
    audiences_sheet = workbook['Audiences']
    row = 2
    for audience in my_audiences:
        audiences_sheet[f'A{row}'] = audience['name']
        audiences_sheet[f'B{row}'] = audience['subject']
        row += 1


def fill_lessons_sheet(workbook, schedule):
    subjects_sheet = workbook['Subjects']
    teachers_sheet = workbook['Teachers']
    audiences_sheet = workbook['Audiences']
    lessons_sheet = workbook['Lessons']
    row = 3
    for lesson in schedule:
        lessons_sheet[f'A{row}'] = lesson['date']['start']
        lessons_sheet[f'B{row}'] = lesson['date']['end']
        lessons_sheet[f'C{row}'] = lesson['time']['start']
        lessons_sheet[f'D{row}'] = lesson['time']['end']

        subject_cell_row = find_row_by_value(subjects_sheet, 'A', lesson['subject'])[0]
        lessons_sheet[f'E{row}'] = f'=IF(E2=TRUE, IFERROR(Subjects!A{subject_cell_row}, ""), "")'
        lessons_sheet[f'F{row}'] = f'=IF(F2=TRUE, IF((IFERROR(Subjects!B{subject_cell_row}, ""))="", "", (IFERROR(Subjects!B{subject_cell_row}, ""))), "")'

        audiences = set()
        for audience in lesson['audiences']:
            subject_cell_rows = find_row_by_value(audiences_sheet, 'A', audience['name'])
            for cell_row in subject_cell_rows:
                if audiences_sheet[f'B{cell_row}'].value == lesson['subject']:
                    audiences.add(cell_row)
        lessons_sheet[f'G{row}'] = f'=IF(G2=TRUE, IFERROR({"&" "&".join(f"Audiences!A{audience}" for audience in audiences)}, ""), "")'

        teachers = set()
        for teacher in lesson['teachers']:
            teacher_cell_row = find_row_by_value(teachers_sheet, 'A', teacher['name'])[0]
            teachers.add(teacher_cell_row)
        lessons_sheet[f'H{row}'] = f'=IF(H2=TRUE, IFERROR({"&" "&".join(f"Teachers!A{teacher}" for teacher in teachers)}, ""), "")'
        lessons_sheet[f'I{row}'] = f'=IF(I2=TRUE, IFERROR({"&" "&".join(f"Teachers!B{teacher}" for teacher in teachers)}, ""), "")'
        lessons_sheet[f'J{row}'] = f'=IF(J2=TRUE, IF((IFERROR({"&" "&".join(f"Teachers!C{teacher}" for teacher in teachers)}, ""))="", "", (IFERROR({"&" "&".join(f"Teachers!C{teacher}" for teacher in teachers)}, ""))), "")'

        lessons_sheet[f'K{row}'] = f'=IF(K2=TRUE, "{lesson['type']}", "")'
        lessons_sheet[f'L{row}'] = '=IF(L2=TRUE, "Оффлайн", "")' 
        lessons_sheet[f'M{row}'] = '=IF(M2=TRUE, "Обязательно", "")' 

        lessons_sheet[f'O{row}'] = f'=_xlfn.TEXTJOIN(" ", True, E{row}:M{row})'

        row += 1


def find_row_by_value(sheet, column, target_value):
    rows = list()
    for cell in sheet[column]:
        if cell.value == target_value:
            rows.append(cell.row)
    if not rows:
        raise ValueError(f"Ни одна ячейка со значение {target_value} не найдена")
    return rows


def load_custom_schedule():
    file_path = 'output/schedule.xlsx'
    workbook = load_workbook(file_path, data_only=True)
    lessons_sheet = workbook['Lessons']

    schedule = []
    row = 3
    while True:
        date_start = lessons_sheet[f'A{row}'].value
        date_end = lessons_sheet[f'B{row}'].value
        time_start = lessons_sheet[f'C{row}'].value
        time_end = lessons_sheet[f'D{row}'].value
        view = lessons_sheet[f'O{row}'].value

        if time_start is None:
            break

        lesson = {
            'date': {'start': date_start, 'end': date_end},
            'time': {'start': time_start, 'end': time_end},
            'view': view
        }

        schedule.append(lesson)
        row += 1

    return schedule